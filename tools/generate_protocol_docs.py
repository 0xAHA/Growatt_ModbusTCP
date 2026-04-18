#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate Growatt protocol reference documentation pages for MkDocs.

Outputs three markdown files:
  docs/developer/protocol-v139.md      -  Modbus RTU Protocol II V1.39
  docs/developer/protocol-vpp.md       -  VPP Protocol (V2.01/V2.03)
  docs/developer/protocol-offgrid.md   -  Off-Grid Protocol V0.26

Source files:
  Protocols/Growatt_WIT-Modbus-RTU-Protocol-II-V1.39.xlsx
  Protocols/GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx
  Protocols/OffGrid-Modbus-RS485-RS232-RTU-Protocol-V0-26.xlsx
  Protocols/growatt_vpp_protocol_v2.01_registers.csv
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

REPO_ROOT = Path(__file__).parent.parent
PROTOCOLS_DIR = REPO_ROOT / "Protocols"
DOCS_DIR = REPO_ROOT / "docs" / "developer"

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def clean(value):
    """Normalise a cell value to a stripped string, or empty string if None."""
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def escape_md(text: str) -> str:
    """Escape characters that break Markdown table cells."""
    return text.replace("|", "\\|").replace("\n", " ")


def md_row(*cells) -> str:
    return "| " + " | ".join(escape_md(str(c)) for c in cells) + " |"


# ---------------------------------------------------------------------------
# V1.39  -  Modbus RTU Protocol II
# ---------------------------------------------------------------------------

V139_EXCEL = PROTOCOLS_DIR / "Growatt_WIT-Modbus-RTU-Protocol-II-V1.39.xlsx"

# Verified bounds (1-indexed rows)
V139_HOLDING_START = 37
V139_HOLDING_END   = 862
V139_INPUT_START   = 864
V139_INPUT_END     = 2004

COL_V139_ADDR   = 1   # int = register address; str = section marker
COL_V139_NAME   = 4
COL_V139_DESC   = 8
COL_V139_ACCESS = 20

SECTION_KEYWORDS = (
    "group", "use for", "for tl", "for wit", "for sph", "for spa",
    "note", "register table", "appendix", "bms", "apx", "reserved",
    "second", "third", "fourth", "fifth", "sixth", "seventh", "eighth",
    "ninth", "tenth", "eleventh", "twelfth", "thirteenth",
)


def _is_v139_section_marker(value) -> bool:
    if not isinstance(value, str):
        return False
    vl = value.lower().strip()
    return any(k in vl for k in SECTION_KEYWORDS)


def _parse_v139_section(rows, start_row, end_row):
    """
    Yield dicts: {address, name, desc, access, section}.
    rows is a 0-indexed list matching 1-indexed spreadsheet rows.
    Section markers become the 'section' label for subsequent registers.
    """
    current_section = None
    current_reg = None

    def emit(reg):
        if reg:
            yield reg

    for sheet_row in range(start_row, end_row + 1):
        row = rows[sheet_row - 1]

        addr_val   = row[COL_V139_ADDR   - 1] if len(row) >= COL_V139_ADDR   else None
        name_val   = row[COL_V139_NAME   - 1] if len(row) >= COL_V139_NAME   else None
        desc_val   = row[COL_V139_DESC   - 1] if len(row) >= COL_V139_DESC   else None
        access_val = row[COL_V139_ACCESS - 1] if len(row) >= COL_V139_ACCESS else None

        if isinstance(addr_val, str) and _is_v139_section_marker(addr_val):
            if current_reg:
                yield current_reg
                current_reg = None
            current_section = clean(addr_val)
            continue

        if isinstance(addr_val, (int, float)) and 0 <= int(addr_val) <= 70000:
            if current_reg:
                yield current_reg
            addr_int = int(addr_val)
            current_reg = {
                "address": addr_int,
                "name": clean(name_val) or f"reg_{addr_int}",
                "desc": clean(desc_val) or "",
                "access": clean(access_val) or "",
                "section": current_section or "",
            }
        elif current_reg:
            # Continuation row  -  append to description
            for col_val in (desc_val, name_val):
                if col_val is not None:
                    extra = clean(col_val)
                    if extra:
                        current_reg["desc"] = (current_reg["desc"] + " " + extra).strip()
                        break

    if current_reg:
        yield current_reg


def _registers_to_md_table(registers, has_range=False):
    """Render a list of register dicts as a markdown table string."""
    lines = []
    if has_range:
        lines.append("| Address | Name | Description | Access | Range |")
        lines.append("| --- | --- | --- | --- | --- |")
    else:
        lines.append("| Address | Name | Description | Access |")
        lines.append("| --- | --- | --- | --- |")

    last_section = None
    for reg in registers:
        sec = reg.get("section", "")
        if sec and sec != last_section:
            # Section divider as italic row
            if has_range:
                lines.append(md_row(f"*{sec}*", "", "", "", ""))
            else:
                lines.append(md_row(f"*{sec}*", "", "", ""))
            last_section = sec

        if has_range:
            lines.append(md_row(
                reg["address"],
                reg["name"],
                reg["desc"],
                reg["access"],
                reg.get("range", ""),
            ))
        else:
            lines.append(md_row(
                reg["address"],
                reg["name"],
                reg["desc"],
                reg["access"],
            ))

    return "\n".join(lines)


def generate_v139():
    print(f"Loading {V139_EXCEL.name}...")
    wb = openpyxl.load_workbook(V139_EXCEL, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(max_row=V139_INPUT_END + 5, values_only=True))
    wb.close()
    print(f"  {len(rows)} rows loaded")

    print("  Parsing holding registers...")
    holding = list(_parse_v139_section(rows, V139_HOLDING_START, V139_HOLDING_END))
    print(f"  {len(holding)} holding registers")

    print("  Parsing input registers...")
    inputs = list(_parse_v139_section(rows, V139_INPUT_START, V139_INPUT_END))
    print(f"  {len(inputs)} input registers")

    out = DOCS_DIR / "protocol-v139.md"

    content = f"""\
# Modbus RTU Protocol II V1.39

> **Source document:** Growatt WIT Modbus RTU Protocol II V1.39
> (`Growatt_WIT-Modbus-RTU-Protocol-II-V1.39.xlsx`)
>
> This is the definitive Protocol II document. V1.39 is a verified superset of
> V1.13, V1.20, and V1.24.
>
> **Applicable models:** WIT, WIS, SPH, SPA, MIN, MOD, MID, MAX, MAC

---

## Register Ranges by Model Family

| Family | Register Ranges |
| --- | --- |
| TL-X / TL-XH (MIN type) | 0-124, 3000-3124, 3125-3249, 3250-3374 |
| TL3-X (MAX / MID / MAC) | 0-124, 3000-3124, 3125-3249 |
| SPA / SPH (hybrid) | 0-124, 1000-1124, 2000-2124, 3000-3124, 3125-3249 |
| WIT TL3 | 0-124, 875-999, 3000-3124, 3125-3249, 8000-8139 |

---

## Holding Registers ({len(holding)} registers)

{_registers_to_md_table(holding)}

---

## Input Registers ({len(inputs)} registers)

{_registers_to_md_table(inputs)}
"""

    out.write_text(content, encoding="utf-8")
    print(f"  Written: {out}")


# ---------------------------------------------------------------------------
# VPP Protocol
# ---------------------------------------------------------------------------

VPP_CSV   = PROTOCOLS_DIR / "growatt_vpp_protocol_v2.01_registers.csv"
VPP_EXCEL = PROTOCOLS_DIR / "GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx"


def _load_vpp_registers():
    """Load Hold and Input registers from the VPP V2.01 CSV."""
    holding = []
    inputs  = []
    with open(VPP_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rt = row["Register_Type"].strip()
            try:
                addr = int(row["Address"].strip())
            except (ValueError, KeyError):
                continue
            rec = {
                "address":  addr,
                "name":     clean(row.get("Parameter_Name", "")),
                "access":   clean(row.get("Read_Write", "")),
                "dtype":    clean(row.get("Data_Type", "")),
                "unit":     clean(row.get("Unit", "")),
                "count":    clean(row.get("Register_Count", "")),
                "range":    clean(row.get("Range_Notes", "")),
            }
            if rt == "Hold":
                holding.append(rec)
            elif rt == "Input":
                inputs.append(rec)
    return holding, inputs


def _load_vpp_dtc_codes():
    """Extract DTC codes from the VPP CSV reference table."""
    dtc_codes = []
    in_dtc = False
    with open(VPP_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rt = row["Register_Type"].strip()
            no = row["NO"].strip()
            if "TABLE 3-1" in rt.upper():
                in_dtc = True
                continue
            if in_dtc:
                if rt.startswith("TABLE 3-2") or rt.startswith("TABLE 3-3") or rt.startswith("TABLE 3-4"):
                    break
                if rt == "Model" or rt == "":
                    continue
                try:
                    dtc = int(no)
                    dtc_codes.append({"model": rt, "dtc": dtc})
                except ValueError:
                    pass
    return dtc_codes


def _vpp_registers_to_md_table(registers):
    lines = [
        "| Address | Parameter Name | R/W | Type | Unit | Count | Notes |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for reg in registers:
        lines.append(md_row(
            reg["address"],
            reg["name"],
            reg["access"],
            reg["dtype"],
            reg["unit"],
            reg["count"],
            reg["range"],
        ))
    return "\n".join(lines)


def generate_vpp():
    print(f"Loading VPP registers from {VPP_CSV.name}...")
    holding, inputs = _load_vpp_registers()
    print(f"  {len(holding)} holding, {len(inputs)} input registers")

    print("  Loading DTC codes...")
    dtc_codes = _load_vpp_dtc_codes()
    print(f"  {len(dtc_codes)} DTC codes")

    dtc_rows = "\n".join(
        f"| {d['dtc']} | {escape_md(d['model'])} |"
        for d in sorted(dtc_codes, key=lambda x: x["dtc"])
    )

    out = DOCS_DIR / "protocol-vpp.md"

    content = f"""\
# VPP Protocol

> **Source documents:** Growatt VPP Communication Protocol of Inverter V2.01 / V2.03
> (`GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx`,
> `growatt_vpp_protocol_v2.01_registers.csv`)
>
> VPP (Virtual Power Plant) is Growatt's advanced monitoring and control protocol.
> It uses registers in the 30000+ range and is only supported by newer inverter models.
> Multiple protocol versions (V2.01, V2.02, V2.03) are in active use  -  all share the
> same core register layout described here.
>
> **Applicable models:** SPH, SPA, MIN TL-XH, MOD TL3-XH, MID TL3-XH, WIT, WIS, and others

---

## Register Ranges

| Range | Purpose |
| --- | --- |
| 30000-30099 | Device identification, rated parameters, system settings |
| 30100-30499 | Control registers (AC power, battery, TOU schedule) |
| 31000-31499 | Real-time data (status, PV, grid, battery, load) |
| 32000+ | Extended / model-specific |

---

## DTC Codes (Table 3-1)

The DTC (Device Type Code) is stored at holding register 30000 and uniquely identifies
the inverter model. The integration reads this at startup for automatic model detection.

| DTC Code | Model |
| --- | --- |
{dtc_rows}

---

## Holding Registers ({len(holding)} registers)

{_vpp_registers_to_md_table(holding)}

---

## Input Registers ({len(inputs)} registers)

{_vpp_registers_to_md_table(inputs)}
"""

    out.write_text(content, encoding="utf-8")
    print(f"  Written: {out}")


# ---------------------------------------------------------------------------
# Off-Grid Protocol V0.26
# ---------------------------------------------------------------------------

OFFGRID_EXCEL = PROTOCOLS_DIR / "OffGrid-Modbus-RS485-RS232-RTU-Protocol-V0-26.xlsx"

OFFGRID_HOLDING_START = 56
OFFGRID_HOLDING_END   = 436

COL_OG_ADDR   = 1
COL_OG_NAME   = 7
COL_OG_DESC   = 21
COL_OG_ACCESS = 33
COL_OG_RANGE  = 37

OG_SECTION_KEYWORDS = (
    "group", "use for", "note", "register table", "bms", "battery",
    "ac output", "ac input", "solar", "utility", "second", "third",
    "fourth", "fifth", "sixth", "fault", "warning", "status",
    "set address", "reserved",
)

# Input registers sourced directly from the Off-Grid Protocol V0.26 PDF.
# The Excel input section uses an inconsistent multi-column layout with merged
# cells across multiple registers, making reliable automated extraction impossible.
# Format: (address, variable_name, description, scale, unit, notes)
OFFGRID_INPUT_REGISTERS = [
    (0,   "System Status",        "System run state",
     "-", "-",
     "0:Standby, 1:PV&Grid Supporting Loads, 2:Battery Discharging, 3:Fault, 4:Flash, "
     "5:PV Charging, 6:Grid Charging, 7:PV&Grid Charging, 8:PV&Grid Charging+Grid Bypass, "
     "9:PV Charging+Grid Bypass, 10:Grid Charging+Grid Bypass, 11:Grid Bypass, "
     "12:PV Charging+Loads Supporting, 13:PV Discharging, 14:PV&Battery Discharging, "
     "15:Gen Charging, 16:Gen Charging+Gen Bypass, 17:PV&Gen Charging, "
     "18:PV&Gen Charging+Gen Bypass, 19:PV Charging+Gen Bypass, 20:Gen Bypass, "
     "21:PV Export to Grid, 22:PV Export to Grid+Loads Supporting, "
     "23:PV Charging+Export to Grid, 24:PV Charging+Export to Grid+Loads Supporting, "
     "25:Battery Export to Grid, 26:Battery Export to Grid+Loads Supporting, "
     "27:Battery&PV Export to Grid, 28:Battery&PV Export to Grid+Loads Supporting"),
    (1,   "Vpv1",                 "PV1 voltage",                          "0.1",  "V",   ""),
    (2,   "Vpv2",                 "PV2 voltage",                          "0.1",  "V",   ""),
    (3,   "Ppv1 H",               "PV1 charge power (high)",              "0.1",  "W",   ""),
    (4,   "Ppv1 L",               "PV1 charge power (low)",               "0.1",  "W",   ""),
    (5,   "Ppv2 H",               "PV2 charge power (high)",              "0.1",  "W",   ""),
    (6,   "Ppv2 L",               "PV2 charge power (low)",               "0.1",  "W",   ""),
    (7,   "Buck1Curr/Pv1Curr",    "Buck1 current or PV1 current",         "0.1",  "A",   ""),
    (8,   "Buck2Curr/Pv2Curr",    "Buck2 current or PV2 current",         "0.1",  "A",   ""),
    (9,   "OP_Watt H",            "Output active power (high)",           "0.1",  "W",   ""),
    (10,  "OP_Watt L",            "Output active power (low)",            "0.1",  "W",   ""),
    (11,  "OP_VA H",              "Output apparent power (high)",         "0.1",  "VA",  ""),
    (12,  "OP_VA L",              "Output apparent power (low)",          "0.1",  "VA",  ""),
    (13,  "ACChr_Watt H",         "AC charge watt (high)",                "0.1",  "W",   ""),
    (14,  "ACChr_Watt L",         "AC charge watt (low)",                 "0.1",  "W",   ""),
    (15,  "ACChr_VA H",           "AC charge apparent power (high)",      "0.1",  "VA",  ""),
    (16,  "ACChr_VA L",           "AC charge apparent power (low)",       "0.1",  "VA",  ""),
    (17,  "Bat Volt",             "Battery voltage (M3)",                 "0.01", "V",   ""),
    (18,  "BatterySOC",           "Battery SOC",                         "1",    "%",   "0~100"),
    (19,  "Bus Volt",             "INV bus total voltage",                "0.1",  "V",   ""),
    (20,  "Grid Volt",            "AC input voltage",                     "0.1",  "V",   ""),
    (21,  "Line Freq",            "AC input frequency",                   "0.01", "Hz",  ""),
    (22,  "OutputVolt",           "AC output voltage",                    "0.1",  "V",   ""),
    (23,  "OutputFreq",           "AC output frequency",                  "0.01", "Hz",  ""),
    (24,  "Ouput DCV",            "Output DC voltage",                    "0.1",  "V",   ""),
    (25,  "InvTemp",              "Inverter temperature",                 "0.1",  "°C",  "-30~200.0"),
    (26,  "DcDc Temp",            "DC-DC temperature",                    "0.1",  "°C",  "-30~200.0"),
    (27,  "LoadPercent",          "Load percentage",                      "0.1",  "%",   "0~1000"),
    (28,  "Bat_s_Volt",           "Battery-port voltage (DSP)",           "0.01", "V",   ""),
    (29,  "Bat_Volt_DSP",         "Battery-bus voltage (DSP)",            "0.01", "V",   ""),
    (30,  "Time total H",         "Work time total (high)",               "0.5",  "S",   ""),
    (31,  "Time total L",         "Work time total (low)",                "0.5",  "S",   ""),
    (32,  "Buck1_NTC",            "Buck1 temperature",                    "0.1",  "°C",  "-30~200.0"),
    (33,  "Buck2_NTC",            "Buck2 temperature",                    "0.1",  "°C",  "-30~200.0"),
    (34,  "OP_Curr",              "Output current",                       "0.1",  "A",   ""),
    (35,  "Inv_Curr",             "Inverter current",                     "0.1",  "A",   ""),
    (36,  "AC_InWatt H",          "AC input watt (high)",                 "0.1",  "W",   "signed int32; >0: get energy from grid, <0: export to grid"),
    (37,  "AC_InWatt L",          "AC input watt (low)",                  "0.1",  "W",   ""),
    (38,  "AC_InVA H",            "AC input apparent power (high)",       "0.1",  "VA",  ""),
    (39,  "AC_InVA L",            "AC input apparent power (low)",        "0.1",  "VA",  ""),
    (40,  "Fault bit",            "Fault bit",                            "-",    "-",   ""),
    (41,  "Warning bit",          "Warning bit",                          "-",    "-",   ""),
    (42,  "Warning bit high",     "Warning bit high",                     "-",    "-",   ""),
    (43,  "warning value",        "Warning value",                        "-",    "-",   ""),
    (44,  "DTC",                  "Device Type Code",                     "-",    "-",   ""),
    (45,  "Export to Grid Today", "Today's energy fed to grid",           "0.1",  "kWh", ""),
    (46,  "Export to Grid Total H","Total energy fed to grid (high)",     "0.1",  "kWh", ""),
    (47,  "Export to Grid Total L","Total energy fed to grid (low)",      "0.1",  "kWh", ""),
    (48,  "Epv1_today H",         "PV1 energy today (high)",              "0.1",  "kWh", ""),
    (49,  "Epv1_today L",         "PV1 energy today (low)",               "0.1",  "kWh", ""),
    (50,  "Epv1_total H",         "PV1 energy total (high)",              "0.1",  "kWh", ""),
    (51,  "Epv1_total L",         "PV1 energy total (low)",               "0.1",  "kWh", ""),
    (52,  "Epv2_today H",         "PV2 energy today (high)",              "0.1",  "kWh", ""),
    (53,  "Epv2_today L",         "PV2 energy today (low)",               "0.1",  "kWh", ""),
    (54,  "Epv2_total H",         "PV2 energy total (high)",              "0.1",  "kWh", ""),
    (55,  "Epv2_total L",         "PV2 energy total (low)",               "0.1",  "kWh", ""),
    (56,  "Eac_chrToday H",       "AC charge energy today (high)",        "0.1",  "kWh", ""),
    (57,  "Eac_chrToday L",       "AC charge energy today (low)",         "0.1",  "kWh", ""),
    (58,  "Eac_chrTotal H",       "AC charge energy total (high)",        "0.1",  "kWh", ""),
    (59,  "Eac_chrTotal L",       "AC charge energy total (low)",         "0.1",  "kWh", ""),
    (60,  "Ebat_dischrToday H",   "Battery discharge energy today (high)","0.1",  "kWh", ""),
    (61,  "Ebat_dischrToday L",   "Battery discharge energy today (low)", "0.1",  "kWh", ""),
    (62,  "Ebat_dischrTotal H",   "Battery discharge energy total (high)","0.1",  "kWh", ""),
    (63,  "Ebat_dischrTotal L",   "Battery discharge energy total (low)", "0.1",  "kWh", ""),
    (64,  "Eac_dischrToday H",    "AC discharge energy today (high)",     "0.1",  "kWh", ""),
    (65,  "Eac_dischrToday L",    "AC discharge energy today (low)",      "0.1",  "kWh", ""),
    (66,  "Eac_dischrTotal H",    "AC discharge energy total (high)",     "0.1",  "kWh", ""),
    (67,  "Eac_dischrTotal L",    "AC discharge energy total (low)",      "0.1",  "kWh", ""),
    (68,  "ACChrCurr",            "AC charge battery current",            "0.1",  "A",   ""),
    (69,  "AC_DisChrWatt H",      "AC discharge watt (high)",             "0.1",  "W",   ""),
    (70,  "AC_DisChrWatt L",      "AC discharge watt (low)",              "0.1",  "W",   ""),
    (71,  "AC_DisChrVA H",        "AC discharge apparent power (high)",   "0.1",  "VA",  ""),
    (72,  "AC_DisChrVA L",        "AC discharge apparent power (low)",    "0.1",  "VA",  ""),
    (73,  "Bat_DisChrWatt H",     "Battery discharge watt (high)",        "0.1",  "W",   ""),
    (74,  "Bat_DisChrWatt L",     "Battery discharge watt (low)",         "0.1",  "W",   ""),
    (75,  "Bat_DisChrVA H",       "Battery discharge apparent power (high)","0.1","VA",  ""),
    (76,  "Bat_DisChrVA L",       "Battery discharge apparent power (low)","0.1", "VA",  ""),
    (77,  "Bat_Watt H",           "Battery watt (high)",                  "0.1",  "W",   "signed int32; positive=discharge, negative=charge"),
    (78,  "Bat_Watt L",           "Battery watt (low)",                   "0.1",  "W",   ""),
    (79,  "uwSlaveExistCnt",      "Number of parallel slave units",       "-",    "-",   ""),
    (81,  "MpptFanSpeed",         "MPPT charger fan speed",               "1",    "%",   "0~100"),
    (82,  "InvFanSpeed",          "Inverter fan speed",                   "1",    "%",   "0~100"),
    (83,  "TotalChgCur",          "Total charge current",                 "0.1",  "A",   ""),
    (84,  "TotalDisChgCur",       "Total discharge current",              "0.1",  "A",   ""),
    (85,  "Eop_dischrToday_H",    "Output discharge energy today (high)", "0.1",  "kWh", ""),
    (86,  "Eop_dischrToday_L",    "Output discharge energy today (low)",  "0.1",  "kWh", ""),
    (87,  "Eop_dischrTotal_H",    "Output discharge energy total (high)", "0.1",  "kWh", ""),
    (88,  "Eop_dischrTotal_L",    "Output discharge energy total (low)",  "0.1",  "kWh", ""),
    (90,  "ParaChgCurr",          "Parallel system charge current",       "0.1",  "A",   ""),
    (91,  "ParStatus",            "Parallel status",                      "-",    "-",
     "0:New module, 1:Master, 2:Slave (single parallel), "
     "3:Slave1 (three phase R), 4:Slave2 (three phase S), 5:Slave3 (three phase T), "
     "6:Slave4 (two phase R), 7:Slave5 (two phase/120° S), 8:Slave6 (two phase/180° S)"),
    (92,  "EGen_dischrToday_H",   "Generator energy today (high)",        "0.1",  "kWh", ""),
    (93,  "EGen_dischrToday_L",   "Generator energy today (low)",         "0.1",  "kWh", ""),
    (94,  "EGen_dischrTotal_H",   "Generator energy total (high)",        "0.1",  "kWh", ""),
    (95,  "EGen_dischrTotal_L",   "Generator energy total (low)",         "0.1",  "kWh", ""),
    (96,  "EGen_dischrPower",     "Generator power",                      "1",    "W",   ""),
    (97,  "EGen_voltage",         "Generator voltage",                    "0.1",  "V",   ""),
    (98,  "EBatChgToday_H",       "Battery charge energy today (high)",   "0.1",  "kWh", ""),
    (99,  "EBatChgToday_L",       "Battery charge energy today (low)",    "0.1",  "kWh", ""),
    (100, "EBatChgTotal_H",       "Battery charge energy total (high)",   "0.1",  "kWh", ""),
    (101, "EBatChgTotal_L",       "Battery charge energy total (low)",    "0.1",  "kWh", ""),
    (102, "CT_InWatt H",          "CT input watt (high)",                 "0.1",  "W",   ""),
    (103, "CT_InWatt L",          "CT input watt (low)",                  "0.1",  "W",   ""),
    (104, "CtLoadWatt H",         "CT load active power (high)",          "0.1",  "W",   ""),
    (105, "CtLoadWatt L",         "CT load active power (low)",           "0.1",  "W",   ""),
    (106, "CtLoadPer",            "CT load percentage",                   "0.1",  "%",   "0~1000"),
    (107, "TxTemp",               "Transformer temperature",              "0.1",  "°C",  "-30~200.0"),
    (108, "LLCTemp",              "LLC temperature",                      "0.1",  "°C",  "-30~200.0"),
    (109, "LLCBusVolt",           "LLC bus total voltage",                "0.1",  "V",   ""),
    (110, "LLCBatVolt",           "LLC battery voltage",                  "0.01", "V",   ""),
    (111, "EnvTemp",              "Environment temperature",              "0.1",  "°C",  "-30~200.0"),
    # --- BMS registers (200-290) ---
    (200, "BMS_Status",           "BMS status",                           "-",    "-",   "Bit field, see note *9"),
    (201, "BMS_Error_old",        "BMS error (legacy)",                   "-",    "-",   "Bit field, see note *10"),
    (202, "BMS_WarnInfo_old",     "BMS warning info (legacy)",            "-",    "-",   "Bit field, see note *11"),
    (203, "BMS_SOC",              "BMS state of charge",                  "1",    "%",   "1~100"),
    (204, "BMS_BatteryVolt",      "BMS average voltage",                  "0.01", "V",   ""),
    (205, "BMS_BatteryCurr",      "BMS average current",                  "0.1",  "A",   "signed int16; see note *12"),
    (206, "BMS_BatteryTemp",      "BMS average temperature",              "0.1",  "°C",  "signed int16"),
    (207, "BMS_MaxCurrChg",       "BMS maximum charge current",           "0.1",  "A",   ""),
    (208, "BMS_CVolt",            "BMS float charge voltage",             "0.01", "V",   "see note *13"),
    (209, "BMS_BMSInfo",          "BMS board info",                       "-",    "-",   "see note *14"),
    (210, "BMS_PackInfo",         "Battery module info",                  "-",    "-",   "see note *15"),
    (211, "BMS_UsingCap",         "Battery used capacity",                "-",    "-",   ""),
    (212, "BMS_Cell_Volt1",       "Cell voltage 1",                       "0.001","V",   "Individual cell data  -  identifies different battery packs under same BMS"),
    (213, "BMS_Cell_Volt2",       "Cell voltage 2",                       "0.001","V",   ""),
    (214, "BMS_Cell_Volt3",       "Cell voltage 3",                       "0.001","V",   ""),
    (215, "BMS_Cell_Volt4",       "Cell voltage 4",                       "0.001","V",   ""),
    (216, "BMS_Cell_Volt5",       "Cell voltage 5",                       "0.001","V",   ""),
    (217, "BMS_Cell_Volt6",       "Cell voltage 6",                       "0.001","V",   ""),
    (218, "BMS_Cell_Volt7",       "Cell voltage 7",                       "0.001","V",   ""),
    (219, "BMS_Cell_Volt8",       "Cell voltage 8",                       "0.001","V",   ""),
    (220, "BMS_Cell_Volt9",       "Cell voltage 9",                       "0.001","V",   ""),
    (221, "BMS_Cell_Volt10",      "Cell voltage 10",                      "0.001","V",   ""),
    (222, "BMS_Cell_Volt11",      "Cell voltage 11",                      "0.001","V",   ""),
    (223, "BMS_Cell_Volt12",      "Cell voltage 12",                      "0.001","V",   ""),
    (224, "BMS_Cell_Volt13",      "Cell voltage 13",                      "0.001","V",   ""),
    (225, "BMS_Cell_Volt14",      "Cell voltage 14",                      "0.001","V",   ""),
    (226, "BMS_Cell_Volt15",      "Cell voltage 15",                      "0.001","V",   ""),
    (227, "BMS_Cell_Volt16",      "Cell voltage 16",                      "0.001","V",   ""),
    (228, "ModuleID",             "Module ID",                            "-",    "-",   "1~12"),
    (229, "ModuleTotalVolt",      "Module total voltage",                 "0.01", "V",   "signed int16"),
    (230, "ModuleTotalCurrent",   "Module total current",                 "0.1",  "A",   "signed int16"),
    (231, "ModuleSoc",            "Module state of charge",               "1",    "%",   "1~100"),
    (232, "ModuleStatus",         "Module status",                        "-",    "-",   "see note *16"),
    (233, "BatProtect1_2",        "Battery protection flags 1-2",         "-",    "-",   "see note *17"),
    (234, "BatWarnInfo1_2",       "Battery warning flags 1-2",            "-",    "-",   "see note *18"),
    (235, "PackNumber",           "Number of parallel battery packs",     "-",    "-",   "1~254"),
    (236, "BatDePowerReason",     "Battery power derating reason",        "-",    "-",   "see note *19"),
    (237, "SOH",                  "Battery state of health",              "-",    "%",   "Bit0~Bit6: SOH value; Bit7: battery end-of-life warning flag"),
    (238, "GaugeRM",              "Remaining capacity",                   "10",   "mAh", ""),
    (239, "GaugeFCC",             "Full charge capacity (nominal)",       "10",   "mAh", ""),
    (240, "DeltaV",               "Cell voltage delta",                   "1",    "mV",  ""),
    (241, "CycleCount",           "Charge/discharge cycle count",         "-",    "-",   ""),
    (242, "RequestOrBatteryType", "Charge request or battery type",       "-",    "-",   "see note *20"),
    (243, "MaximumCellVoltage",   "Maximum cell voltage",                 "1",    "mV",  ""),
    (244, "MinimumCellVoltage",   "Minimum cell voltage",                 "1",    "mV",  ""),
    (245, "MaxMinCellVoltageNumber","Max/min cell voltage numbers",       "-",    "-",   "Bit0~Bit7: min voltage cell number; Bit8~Bit15: max voltage cell number"),
    (246, "ProtectPackID",        "Faulted battery pack address",         "-",    "-",   ""),
    (247, "ManufacturerName",     "Manufacturer name",                    "-",    "-",   ""),
    (248, "HardwareVersion",      "Hardware version",                     "-",    "-",   "1~9"),
    (249, "SoftwareVersion01",    "Software version (bytes 0-1)",         "-",    "-",   ""),
    (250, "ParallelHightSoftwarVer","Highest software version in parallel system","-","-",""),
    (251, "MaxCellTemp",          "Maximum cell temperature",             "0.1",  "°C",  "signed int16"),
    (252, "MinCellTemp",          "Minimum cell temperature",             "0.1",  "°C",  "signed int16"),
    (253, "MaxMinCellTempSerialNum","Max/min temperature cell numbers",   "-",    "-",   "Bit0~Bit7: MinCellTempNum; Bit8~Bit15: MaxCellTempNum"),
    (254, "MaxMinSOC",            "Max/min SOC",                          "-",    "%",   "0~100; Bit0~Bit7: MinSOC; Bit8~Bit15: MaxSOC"),
    (255, "TotalCellNumber",      "Total cell count",                     "-",    "-",   "1~254"),
    (256, "BatProtect3_4",        "Battery protection flags 3-4",         "-",    "-",   "see note *21"),
    (257, "BatProtect5",          "Battery protection flags 5",           "-",    "-",   "see note *22"),
    (258, "BatWarnInfo3",         "Battery warning flags 3",              "-",    "-",   "see note *23"),
    (259, "UpdateStatus",         "Firmware upgrade status",              "-",    "-",   "Bit0~1: 0=normal, 1=programming, 2=upgrade successful"),
    (260, "SoftwareVersion23",    "Software version (bytes 2-3)",         "-",    "-",   "ASCII encoded"),
    (261, "SoftwareVersion45",    "Software version (bytes 4-5)",         "-",    "-",   "ASCII encoded"),
    (262, "BatSerialNumber_ID",   "Battery serial number ID",             "-",    "-",   ""),
    (263, "BatSerialNumber0_1",   "Battery serial number (chars 0-1)",    "-",    "-",   "ASCII encoded"),
    (264, "BatSerialNumber2_3",   "Battery serial number (chars 2-3)",    "-",    "-",   "ASCII encoded"),
    (265, "BatSerialNumber4_5",   "Battery serial number (chars 4-5)",    "-",    "-",   "ASCII encoded"),
    (266, "BatSerialNumber6_7",   "Battery serial number (chars 6-7)",    "-",    "-",   "ASCII encoded"),
    (267, "BatSerialNumber8_9",   "Battery serial number (chars 8-9)",    "-",    "-",   "ASCII encoded"),
    (268, "BatSerialNumber10_11", "Battery serial number (chars 10-11)",  "-",    "-",   "ASCII encoded"),
    (269, "BatSerialNumber12_13", "Battery serial number (chars 12-13)",  "-",    "-",   "ASCII encoded"),
    (270, "BatSerialNumber14_15", "Battery serial number (chars 14-15)",  "-",    "-",   "ASCII encoded"),
    (271, "BatSerialNumber16_17", "Battery serial number (chars 16-17)",  "-",    "-",   "ASCII encoded"),
    (272, "BatSerialNumber18_19", "Battery serial number (chars 18-19)",  "-",    "-",   "ASCII encoded"),
    (273, "ModuleID2",            "Module ID 2",                          "-",    "-",   "1~12"),
    (274, "Module2MaxVol",        "Module 2 maximum cell voltage",        "0.01", "V",   ""),
    (275, "Module2MimVol",        "Module 2 minimum cell voltage",        "0.01", "V",   ""),
    (276, "Module2MaxTemp",       "Module 2 maximum temperature",         "1",    "°C",  "offset +40"),
    (277, "Module2MimTemp",       "Module 2 minimum temperature",         "1",    "°C",  "offset +40"),
    (278, "DoStatus",             "Output dry contact status",            "-",    "-",   ""),
    (279, "DsgBatNumber",         "Discharge energy statistics - battery ID","1",  "kWh", ""),
    (280, "DsgEnergyKWH_H",       "Discharge energy (high 16 bits)",      "1",    "kWh", ""),
    (281, "DsgEnergyKWH_L",       "Discharge energy (low 16 bits)",       "1",    "kWh", ""),
    (282, "ChgBatNumber",         "Charge energy statistics - battery ID", "-",   "-",   ""),
    (283, "ChgEnergyKWH_H",       "Charge energy (high 16 bits)",         "1",    "kWh", ""),
    (284, "ChgEnergyKWH_L",       "Charge energy (low 16 bits)",          "1",    "kWh", ""),
    (285, "reserve285",           "Reserved",                             "-",    "-",   ""),
    (286, "reserve286",           "Reserved",                             "-",    "-",   ""),
    (287, "reserve287",           "Reserved",                             "-",    "-",   ""),
    (288, "reserve288",           "Reserved",                             "-",    "-",   ""),
    (289, "reserve289",           "Reserved",                             "-",    "-",   ""),
    (290, "reserve290",           "Reserved",                             "-",    "-",   ""),
]


def _offgrid_input_to_md_table():
    lines = [
        "| Address | Variable Name | Description | Scale | Unit | Notes |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for addr, name, desc, scale, unit, notes in OFFGRID_INPUT_REGISTERS:
        lines.append(md_row(addr, name, desc, scale, unit, notes))
    return "\n".join(lines)


def _is_og_section_marker(value) -> bool:
    if not isinstance(value, str):
        return False
    vl = value.lower().strip()
    return any(k in vl for k in OG_SECTION_KEYWORDS)


def _parse_offgrid_section(rows, start_row, end_row):
    """Yield register dicts for one section of the Off-Grid Excel."""
    current_section = None
    current_reg = None

    for sheet_row in range(start_row, end_row + 1):
        row = rows[sheet_row - 1]

        addr_val   = row[COL_OG_ADDR   - 1] if len(row) >= COL_OG_ADDR   else None
        name_val   = row[COL_OG_NAME   - 1] if len(row) >= COL_OG_NAME   else None
        desc_val   = row[COL_OG_DESC   - 1] if len(row) >= COL_OG_DESC   else None
        access_val = row[COL_OG_ACCESS - 1] if len(row) >= COL_OG_ACCESS else None
        range_val  = row[COL_OG_RANGE  - 1] if len(row) >= COL_OG_RANGE  else None

        if isinstance(addr_val, str) and _is_og_section_marker(addr_val):
            if current_reg:
                yield current_reg
                current_reg = None
            current_section = clean(addr_val)
            continue

        if isinstance(addr_val, (int, float)) and 0 <= int(addr_val) <= 70000:
            if current_reg:
                yield current_reg
            addr_int = int(addr_val)
            current_reg = {
                "address": addr_int,
                "name":    clean(name_val) or f"reg_{addr_int}",
                "desc":    clean(desc_val) or "",
                "access":  clean(access_val) or "",
                "range":   clean(range_val) or "",
                "section": current_section or "",
            }
        elif current_reg:
            extra = clean(desc_val) or clean(name_val)
            if extra:
                current_reg["desc"] = (current_reg["desc"] + " " + extra).strip()
            if range_val and not current_reg.get("range"):
                current_reg["range"] = clean(range_val)

    if current_reg:
        yield current_reg


# Reference tables sourced from Off-Grid Protocol V0.26 PDF appendix.
# Chinese descriptions translated to English throughout.
OFFGRID_REFERENCE_TABLES_MD = """\
## Reference Tables

### *0 — System Run State (register 0)

| Value | Status | Description |
| --- | --- | --- |
| 0 | Standby | Standby mode |
| 1 | PV&Grid Supporting Loads | PV and grid combined load support |
| 2 | Battery Discharging | Battery discharging |
| 3 | Fault | Fault condition |
| 4 | Flash | Firmware programming mode (not shown on monitor) |
| 5 | PV Charging | PV charging |
| 6 | Grid Charging | Grid charging |
| 7 | PV&Grid Charging | PV and grid combined charging |
| 8 | PV&Grid Charging+Grid Bypass | Combined charging with grid bypass load |
| 9 | PV Charging+Grid Bypass | PV charging with grid bypass load |
| 10 | Grid Charging+Grid Bypass | Grid charging with bypass load |
| 11 | Grid Bypass | Grid bypass load only |
| 12 | PV Charging+Loads Supporting | PV charging with inverter load support |
| 13 | Export to Grid | Grid-tied export mode |

---

### *1 — Fault Codes (register 40, bit field — see *8 for warning bits)

| Code | Description |
| --- | --- |
| 1 | Fan lock |
| 2 | Over temperature |
| 3 | Battery voltage high |
| 4 | Battery voltage low |
| 5 | Output short circuit |
| 6 | Output voltage high |
| 7 | Overload |
| 8 | DC bus voltage high |
| 9 | DC bus soft-start fail |
| 11 | Main relay fault |
| 51 | Over current |
| 52 | DC bus voltage low |
| 53 | Inverter soft-start fail |
| 56 | IGBT over current |
| 58 | Output voltage low |
| 60 | Negative power excessive |
| 61 | PV voltage high |
| 62 | Internal SCI communication error |
| 80 | CAN communication fault |
| 81 | Master unit lost |

---

### *6 — DTC Code Description (register 44)

| Code Range | Device Type |
| --- | --- |
| 03xxx | PV Storage — front 1 tracker PV Storage |

---

### *8 — Warning Code Bits (registers 41 and 42)

**Warning bit (register 41)**

| Bit mask | Bit | Description |
| --- | --- | --- |
| 0x0001 | 0 | Fan lock warning |
| 0x0002 | 1 | Battery over-charge |
| 0x0004 | 2 | Battery voltage low |
| 0x0008 | 3 | Overload |
| 0x0010 | 4 | Output power derating |
| 0x0020 | 5 | Solar stopped — battery voltage too low |
| 0x0040 | 6 | Solar stopped — PV voltage too high |
| 0x0080 | 7 | Solar stopped — overload |
| 0x0100 | 8 | Parallel grid input inconsistent |
| 0x0200 | 9 | Parallel input phase sequence error |
| 0x0400 | 10 | Parallel output phase loss |
| 0x0800 | 11 | Over temperature |
| 0x1000 | 12 | Buck current excessive |
| 0x2000 | 13 | Battery disconnected |
| 0x4000 | 14 | BMS communication error |
| 0x8000 | 15 | PV power insufficient |

**Warning bit high (register 42)**

| Bit mask | Bit | Description |
| --- | --- | --- |
| 0x0001 | 0 | No battery — parallel disabled |
| 0x0002 | 1 | Parallel firmware version mismatch |
| 0x0004 | 2 | Reserved |
| 0x0008 | 3 | Parallel unit capacity mismatch |
| 0x0010 | 4 | Master unit lost |
| 0x0020 | 5 | BMS cell over-voltage |
| 0x0040 | 6 | BMS total over-voltage |
| 0x0080 | 7 | BMS discharge over-current |
| 0x0100 | 8 | BMS charge over-current |
| 0x0200 | 9 | BMS over-temperature |
| 0x0400 | 10 | Battery voltage inconsistency |

---

### *9 — BMS_Status Code (register 200)

| Bit(s) | Content | Values |
| --- | --- | --- |
| 0-1 | Operating status | 00: soft-starting; 01: standby; 10: charging; 11: discharging |
| 2 | Error byte valid flag | 0: error byte invalid; 1: error byte valid |
| 3 | Cell balance PF status | 0: unbalanced; 1: balanced |
| 4 | Sleep status | 0: disabled; 1: enabled |
| 5 | Output discharge status | 0: disabled; 1: enabled |
| 6 | Output charge status | 0: disabled; 1: enabled |
| 7 | Battery terminal status | 0: terminal connected; 1: terminal open |
| 8-9 | Master box operation mode | 00: stand-alone; 01: parallel; 10: parallel preparation |
| 10-11 | SP status | 00: none; 01: standby; 10: charging; 11: discharging |
| 12 | Force charge request | 0: disabled; 1: enabled |

---

### *10 — BMS_Error_old Code (register 201)

| Bit | Description | Recovery |
| --- | --- | --- |
| 0 | OCD — over-current discharge protection | Unload AND (charging OR DG_ON command) |
| 1 | SCD — short-circuit discharge protection | Unload AND (charging OR DG_ON command) |
| 2 | OV — over-voltage protection | Stop charging AND discharging |
| 3 | UV — under-voltage protection | Unload AND charging |
| 4 | OTD — over-temperature discharge protection | Unload AND temperature drops to 60 C |
| 5 | OTC — over-temperature charge protection | Stop charging OR temperature drops to 50 C |
| 6 | UTD — under-temperature discharge protection | Unload AND temperature rises to -10 C |
| 7 | UTC — under-temperature charge protection | Stop charging OR temperature rises to 0 C |
| 8 | Soft-start fail | — |
| 9 | Permanent fault | — |
| 10 | Delta-V fail | — |
| 11 | OCC — over-current charge protection | Unload AND (discharging OR DG_ON command) |
| 12 | OT — MOS over-temperature protection | MOS temperature drops to rated max |
| 13 | OT — environment over-temperature protection | Environment temperature drops to rated max |
| 14 | UT — environment under-temperature protection | Environment temperature rises to rated min |

---

### *11 — BMS_WarnInfo_old Code (register 202)

| Bit | Warning (bit=1) | Recovery condition |
| --- | --- | --- |
| 0 | Cell over-voltage warning | Discharging or voltage falls below cell OV alarm threshold |
| 1 | Cell under-voltage warning | Charging or voltage rises above cell UV alarm threshold |
| 2 | Total over-voltage warning | Discharging or voltage falls below total OV alarm threshold |
| 3 | Total under-voltage warning | Charging or voltage rises above total UV alarm threshold |
| 4 | Discharge over-current warning | Current falls below discharge OC alarm threshold |
| 5 | Charge over-current warning | Current falls below charge OC alarm threshold |
| 6 | Discharge high-temperature warning | Temperature drops below discharge high-temp alarm |
| 7 | Discharge low-temperature warning | Temperature rises above discharge low-temp alarm |
| 8 | Charge high-temperature warning | Temperature drops below charge high-temp alarm |
| 9 | Charge low-temperature warning | Temperature rises above charge low-temp alarm |
| 10 | MOS high-temperature warning | Temperature drops below MOS high-temp alarm |
| 11 | Environment high-temperature warning | Temperature drops below environment high-temp alarm |
| 12 | Environment low-temperature warning | Temperature rises above environment low-temp alarm |
| 13 | System low-voltage pre-shutdown warning | Total voltage rises above shutdown/lockout threshold |
| 14-15 | Battery type | 00: LiFePO4; 01: Ternary (NMC/NCA); 10: Lithium titanate (LTO); 11: Reserved |

---

### *12 — BMS_BatteryCurr Sign Convention (register 205)

Signed 16-bit integer. Positive current = charging; negative current = discharging.

| Range | Meaning |
| --- | --- |
| 0x0000 to 0x7FFF | Positive current (charging) |
| 0x8000 to 0xFFFF | Negative current (discharging) |

---

### *13 — BMS_CVolt — Float Charge Voltage by Battery Type (register 208)

| Battery Type | CV Voltage |
| --- | --- |
| LiFePO4 (lithium iron phosphate) | 57.6 V |
| Ternary lithium (NMC/NCA) | Set by pack manufacturer |
| Lithium titanate (LTO) | Set by pack manufacturer |

---

### *14 — BMS_BMSInfo Code (register 209)

| Bits | Content |
| --- | --- |
| 0-7 | BMS manufacturer code (00000000 = unspecified; values assigned by manufacturer) |
| 8-15 | BMS generation (00000001 = first generation; 00000010 = second generation; ...) |

---

### *15 — BMS_PackInfo Code (register 210)

| Bits | Content |
| --- | --- |
| 0-7 | Pack manufacturer code (000000001 = EVE; other values assigned by manufacturer) |
| 8-15 | Pack generation (000000001 = first generation; 000000010 = second generation; ...) |

---

### *16 — ModuleStatus Code (register 232)

| Bit(s) | Content | Values |
| --- | --- | --- |
| 0-1 | Operating status | 00: soft-starting; 01: standby; 10: charging; 11: discharging |
| 2 | Error byte valid flag | 0: error byte invalid; 1: error byte valid |
| 3 | Cell balance status | 0: unbalanced; 1: balanced |
| 4 | Sleep status | 0: disabled; 1: enabled |
| 5 | Output discharge status | 0: disabled; 1: enabled |
| 6 | Output charge status | 0: disabled; 1: enabled |
| 7 | Battery terminal status | 0: terminal connected; 1: terminal open |
| 8-9 | Master box operation mode | 00: stand-alone; 01: parallel; 10: parallel preparation |
| 10 | Pre-output discharge status | 0: disabled; 1: enabled |
| 11 | Pre-output charge status | 0: disabled; 1: enabled |
| 12-15 | Reserved | — |

---

### *17 — BatProtect1_2 Code (register 233)

| Bit | Description |
| --- | --- |
| 0 | Soft-start fail |
| 1 | Module under-voltage |
| 2 | Module over-voltage |
| 3 | Cell under-voltage |
| 4 | Cell over-voltage |
| 5 | SCD — short-circuit discharge protection |
| 6 | Charge over-current |
| 7 | Discharge over-current 1 |
| 8 | Parallel firmware version mismatch |
| 9 | Parallel fail |
| 10 | Delta-V fail (internal/external voltage difference too large) |
| 11 | MOS control fail |
| 12 | UTC — under-temperature charge protection |
| 13 | UTD — under-temperature discharge protection |
| 14 | OTC — over-temperature charge protection |
| 15 | OTD — over-temperature discharge protection |

---

### *18 — BatWarnInfo1_2 Code (register 234)

| Bit | Description |
| --- | --- |
| 0 | SOC low warning |
| 1 | Module under-voltage warning |
| 2 | Module over-voltage warning |
| 3 | Cell under-voltage warning |
| 4 | Cell over-voltage warning |
| 5 | Pre-shutdown warning |
| 6 | Charge over-current warning |
| 7 | Discharge over-current warning 1 |
| 8 | Internal communication fail |
| 9 | External CAN communication fail |
| 10 | Delta-V fail (internal/external voltage difference too large) |
| 11 | External RS485 communication fail |
| 12 | UTC — under-temperature charge warning |
| 13 | UTD — under-temperature discharge warning |
| 14 | OTC — over-temperature charge warning |
| 15 | OTD — over-temperature discharge warning |

---

### *19 — BatDePowerReason Code (register 236)

| Bit | Description |
| --- | --- |
| 0 | Cell voltage high — current limit |
| 1 | Cell voltage low — current limit |
| 2 | Cell temperature high — current limit |
| 3 | Cell temperature low — current limit |
| 4 | Total voltage high — current limit |
| 5 | Total voltage low — current limit |
| 6 | Cell voltage delta — current limit |
| 7 | Temperature delta — current limit |
| 8 | Hardware fault — current limit |
| 9 | Fully charged — current limit |
| 10 | MOS temperature high — current limit |
| 11 | Environment temperature high — current limit |
| 12 | Pre-charge fault — current limit |
| 13 | Communication fault — current limit |
| 14 | Bus fault — current limit |
| 15 | Reserved |

---

### *20 — RequestOrBatteryType Code (register 242)

| Bit(s) | Description |
| --- | --- |
| 0-1 | Battery type: 00=LiFePO4; 01=Ternary (NMC/NCA); 10=Lithium titanate (LTO); 11=Reserved |
| 2-3 | Reserved |
| 4 | Force charge request II |
| 5 | Force charge request I |
| 6 | Discharge enable |
| 7 | Charge enable |

---

### *21 — BatProtect3_4 Code (register 256)

| Bit | Description |
| --- | --- |
| 0 | Charge over-power |
| 1 | Discharge over-power |
| 2 | Parallel duplicate address fault |
| 3 | Pre-charge fail |
| 4 | Pre-charge short circuit |
| 5 | AFE-MCU communication error |
| 6 | Cell failure (FLT_CELL_LOST) |
| 7 | Cell temperature sensor failure (FLT_CELL_TEMP_LOST) |
| 8 | Total voltage sampling fault (FLT_SP_UMAIN) |
| 9 | Temperature short circuit (FLT_TEMP_SC) |
| 10 | Load-side total voltage sampling fault (FLT_SP_ULOAD) |
| 11 | Calibration parameter load fault (FLT_EEP_PARAM) |
| 12 | Hardware over-voltage — AFE OV (FLT_OVP) |
| 13 | Hardware under-voltage — AFE UV (FLT_UVP) |
| 14 | Hardware over-current protection (FLT_OCP) |
| 15 | Hardware discharge over-current fault (FLT_DIS_OCP) |

---

### *22 — BatProtect5 Code (register 257)

| Bit | Description |
| --- | --- |
| 0 | Parallel slave/master battery voltage difference too large (FLT_PRLL_UDIFF_OVER) |
| 1 | Charge current limiting fail (FLT_CH_ILIMIT_NORSP) |
| 2 | Discharge current limiting fail (FLT_DI_ILIMIT_NORSP) |
| 3 | Main circuit open-circuit fault (FLT_BUS_OPEN) |
| 4 | Discharge over-current 2 |
| 5 | MOS temperature high |
| 6 | Cell voltage delta too large |
| 7 | Cell temperature delta too large |

---

### *23 — BatWarnInfo3 Code (register 258)

| Bit | Description |
| --- | --- |
| 0 | Charge over-power warning |
| 1 | Discharge over-power warning |
| 2 | Parallel internal charge circulating current over-current warning |
| 3 | Parallel internal discharge circulating current over-current warning |
| 4 | MOS temperature high warning |
| 5 | Cell voltage delta too large |
| 6 | Cell temperature delta too large |
| 7 | Fully charged |
"""


def generate_offgrid():
    print(f"Loading {OFFGRID_EXCEL.name}...")
    wb = openpyxl.load_workbook(OFFGRID_EXCEL, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(max_row=OFFGRID_HOLDING_END + 5, values_only=True))
    wb.close()
    print(f"  {len(rows)} rows loaded")

    print("  Parsing holding registers...")
    holding = list(_parse_offgrid_section(rows, OFFGRID_HOLDING_START, OFFGRID_HOLDING_END))
    print(f"  {len(holding)} holding registers")

    # Input registers: the Excel input section uses inconsistent multi-column layout
    # with merged cells. Data sourced from OFFGRID_INPUT_REGISTERS (PDF-verified).
    n_input = len(OFFGRID_INPUT_REGISTERS)
    print(f"  {n_input} input registers (from PDF source data)")

    out = DOCS_DIR / "protocol-offgrid.md"

    content = f"""\
# Off-Grid Protocol V0.26

> **Source document:** Growatt OffGrid Modbus RS485/RS232 RTU Protocol V0.26
> (`OffGrid-Modbus-RS485-RS232-RTU-Protocol-V0-26.xlsx`)
>
> This is the protocol used by Growatt's off-grid and AC-coupled inverter families.
> Holding registers cover settings and control; input registers cover real-time data.
>
> **Applicable models:**
> - **SPF 3000-6000 ES PLUS**  -  off-grid with battery, no grid export
> - **SPE 8000-12000 ES**  -  higher-capacity off-grid / AC-coupled storage

---

## Register Ranges

| Range | Purpose |
| --- | --- |
| 0-426 | Holding registers (settings and control) |
| 0-111 | Input registers (real-time data) |

---

## Holding Registers ({len(holding)} registers)

{_registers_to_md_table(holding, has_range=True)}

---

## Input Registers ({n_input} registers)

> Source: Off-Grid Protocol V0.26 PDF (register data extracted directly  -  the Excel
> input register section uses an inconsistent layout with merged cells).

{_offgrid_input_to_md_table()}

---

{OFFGRID_REFERENCE_TABLES_MD}"""

    out.write_text(content, encoding="utf-8")
    print(f"  Written: {out}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    DOCS_DIR.mkdir(parents=True, exist_ok=True)

    missing = [p for p in (V139_EXCEL, VPP_CSV, OFFGRID_EXCEL) if not p.exists()]
    if missing:
        for p in missing:
            print(f"ERROR: not found: {p}")
        sys.exit(1)

    print("\n=== Modbus RTU V1.39 ===")
    generate_v139()

    print("\n=== VPP Protocol ===")
    generate_vpp()

    print("\n=== Off-Grid V0.26 ===")
    generate_offgrid()

    print("\nDone. Three pages written to docs/developer/")


if __name__ == "__main__":
    main()
